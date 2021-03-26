import pandas as pd
from datetime import datetime
import openpyxl 
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook,load_workbook
from openpyxl.chart import (LineChart,ScatterChart,Series, Reference)
from openpyxl.chart.axis import DateAxis

# --------------------------
# -- Creates a Line Chart --
# --------------------------
def line_chart(sheetName, fileName, chartPlacement, xAxisCol, dataCols, numDataPoints, chartTitle, yAxisTitle, percentageCols):
    # Open workbook and select sheet to place chart
    wb = load_workbook(fileName)
    ws = wb[sheetName]
    
    # Create Line Chart and set styling
    chart = LineChart()
    chart.title = chartTitle
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None
    chart.y_axis.title = yAxisTitle
    chart.x_axis.tickLblPos = "low"
    chart.legend.position = 'b'
    chart.height = 10
    chart.width = 20
    
    # Set Y-axis Data
    data = Reference(ws, min_col=dataCols[0], min_row=1, max_col=dataCols[0], max_row=numDataPoints + 2)
    chart.add_data(data, titles_from_data=True)
    
    # Set X-axis Data
    dates = Reference(ws, min_col=xAxisCol, min_row=2, max_row=numDataPoints + 2)
    chart.set_categories(dates)
    
    # Converting columns to type % if needed
    for col in percentageCols:
        for i in range(2,numDataPoints+2):
            ws[col + str(i)].number_format = '0.00%'

    # Add the chart to the sheet and save workbook
    ws.add_chart(chart, chartPlacement)
    wb.save(fileName)
    
   
# -----------------------------
# -- Creates a Scatter Chart --
# -----------------------------
def scatter_chart(sheetName, fileName, chartPlacement,dateCol, dataCols, numDataPoints, chartTitle, yAxisTitle, percentageCols):
    # Open workbook and select sheet to place chart
    wb = load_workbook(fileName)
    ws = wb[sheetName]

    # Create Scatter Chart and set styling
    chart = ScatterChart()
    chart.title = chartTitle
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None
    chart.y_axis.title = yAxisTitle
    chart.x_axis.tickLblPos = "low"
    chart.legend.position = 'b'
    chart.height = 10
    chart.width = 20
    
    # Create series
    xvalues = Reference(ws, min_col=1, min_row=2, max_row=numDataPoints+2)
    for i in range(2, 4):
        values = Reference(ws, min_col=i, min_row=1, max_row=numDataPoints+2)
        series = Series(values, xvalues, title_from_data=True)
        chart.series.append(series)
        
    # Converting columns to type % if needed
    for col in percentageCols:
        for i in range(2,numDataPoints+2):
            ws[col + str(i)].number_format = '0.00%'
     
    # Add the chart to the sheet and save workbook           
    ws.add_chart(chart, chartPlacement)
    wb.save(fileName)
